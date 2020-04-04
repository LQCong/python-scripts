# coding:utf-8
import xlrd
import openpyxl
import datetime
import logging
import os
import csv
import re
from functools import wraps


logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)  # Log等级总开关
# logging.basicConfig(level=logging.info, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(processName)s - [ %(levelname)s ]：  %(message)s')


def run_time(func):
    """运行耗时装饰器"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        stime = datetime.datetime.now()
        logger.info("开始执行")
        res = func(*args, **kwargs)
        etime = datetime.datetime.now()
        cost = str(etime - stime).split(":")
        logger.info("执行耗时：[ " + cost[0] + '时' + cost[1] + "分" + cost[2][0:5] + "秒" + " ]")
        return res
    return wrapper


def open_workbook(file_path):
    logger.info("打开文件 %s " % file_path)
    if file_path.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file_path)
        return wb
    elif file_path.endswith('.xls'):
        wb = xlrd.open_workbook(file_path)
        return wb
    else:
        raise TypeError


def is_xlsx(workbook: object or str)->bool:
    if isinstance(workbook, str):
        return True if workbook.endswith('.xlsx') else False
    else:
        return True if str(type(workbook)).split("'")[1].split(".")[0] == 'openpyxl' else False


def get_sheets(workbook: object, sheet_name=None) -> list or object:
    """
        获取工作表中所有工作簿，返回工作簿对象列表
        workbook: 工作表对象
        sheet_name: 获取指定名称的sheet
        :return 返回工作表中的非空工作簿的[ 对象 ]列表, 或返回指定名称的sheet对象
    """
    logger.debug("获取 %s 表" % sheet_name)
    if is_xlsx(workbook):
        ws = workbook[sheet_name]
        return ws
    else:
        ws = workbook.sheet_by_name(sheet_name)
        return ws

def get_sheet_rows_and_cols(sheet: object):
    """
        获取工作簿的行数和列数，返回工作簿对象的行、列的数组
        sheet: 工作簿sheet对象
        :return 返回工作簿行数和列数
    """
    try:
        sheet_type = True if str(type(sheet)).split("'")[1].split(".")[0] == 'openpyxl' else False
        if sheet_type:
            return sheet.max_row, sheet.max_column
        else:
            return sheet.nrows, sheet.ncols
    except AttributeError as e:
        logger.error("获取行数、列数错误，参数错误")
        raise e


def get_table_b_names(workbook: object, table_b_index_column=2)->set:
    """
        获取工作表中所有工作簿的的匹配名称列表，及计数
        workbook: 工作表(非工作簿sheet)对象
        table_b_index_column: 表B索引列，数值从1开始
        :return 返回索引列的计数，格式{'name1': count, 'name2':count,....}
    """
    b_names = set()
    if is_xlsx(workbook):
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            rows, cols = get_sheet_rows_and_cols(ws)
            if not rows == 0:
                for row_index in range(1, rows + 1):
                    if ws.cell(row=row_index, column=table_b_index_column).value:
                        b_names.add(ws.cell(row=row_index, column=table_b_index_column).value)

    else:
        for sheet in workbook.sheet_names():
            ws = workbook.sheet_by_name(sheet)
            rows, cols = get_sheet_rows_and_cols(ws)
            if not ws.nrows == 0:
                for row_index in range(rows):
                    if ws.cell(row_index, table_b_index_column - 1).value:
                        b_names.add(ws.cell(row_index, table_b_index_column - 1).value)
    try:
        workbook.close()
    except:
        pass

    return b_names


def save_work_book(workbook: object, save_file_path: str):
    """保存工作表"""
    logger.info("开始保存数据")
    workbook.save(save_file_path)
    logger.info("保存数据成功")


def get_cell_header(sheet, write_file):
    """获取excel中所有表头"""
    rows, cols = get_sheet_rows_and_cols(sheet)

    string = '姓名|性名|姓名*|员工姓名|姓名▼|会员|户主|姓名▼|員工姓名'

    pattern = re.compile(string)

    for i in range(1, rows + 1):
        li = []
        status = False
        # print(value, type(value))
        # if not (value == "None" or value == "" or value is None):
        for j in range(1, cols+1):
            value = str(sheet.cell(i, j).value)
            li.append(value)

            if value:
                s = re.search(pattern, value)

                if s and s.group():
                    status = True

        if i % 10:
            print(i)

        if status:
            try:
                print(li)
                write_file.write(','.join(li) + '\n')
                # print(li)
            except:
                pass


def clear_content(ws, table_b_names, ws_name):
    """清洗excel内容"""
    # 表头字段汇总
    titles = {'姓名', '招聘渠道', '资源渠道', '面试渠道', '报到渠道', '招募渠道', '入職渠道', '渠道', '原来源渠道', '中介名称',
              '派遣公司A（面试渠道）', '派遣公司B（报到渠道）', '中介字段', '中介来源', '中介栏位', '供应商', '来源',
              '返费', '返费金额', '实返费', '员工返费金额', '员工返费-厂商提供', '打款员工返费金额', '打款员工金额', '中介', '应返费用',
              '报到返费', '返费价格', '报导返费', '返费情况', '打款员工返费', '应发返费', '000返费', '第一笔返费', '员工返费',
              '在职打卡25天返费', '18天返费', '25天返费', '30天返费', '45天返费', '应付返费金额', '报道返费价格',
              '18天返', '返费（给求职者）', '供应商', '价格（报到返）', '供应•商', '业务所属', '供鹿商', '供应豚', '棋应商',
              '费用', '报到返费  到职日期  离职日期 离职原因', '員工姓名', '招聘渠道', '返费价格', '应支付费用', '价格', '实际返钱',
              '结帐金额', '受款厂商', '原供应商', '管道', '利润', '入职反费', '25天返', '返款金额', '一次返费', '二次返费', '报道返费',
              '18天', '按照朱巍巍价格', '报到价格', '报到返 金额', '实际 返钱', '定价', '联系人', '派遣公司A', '應出賬（給面試渠道）',
              '應出賬', '后期返费', '33天费用', '姓名*', '金額', '派遣別', '勞務公司', '勞務公司名稱', '会员', '派遗公司', '姓名 ▼',
              '姓名▼', '費用', '入职返', '返款明细', '企业返款明细', '户主', '總計', '员工姓名', '派遣公司名稱', '总价',
              '派遣公司', '来源', '受款廠商', '原派遣公司', '派遣公司A（面試）', '金额', '來源供應商', '管理費', '派遣公司B', '與二包商確認金額',
              '实际返金', '实际 返銭', '来源渠道', '介绍人', '企业'
              }

    # {表头字段: CSV保存行字段位置}
    list_index_site = {'姓名': 0, '招聘渠道': 1, '资源渠道': 1, '面试渠道': 1, '报到渠道': 1, '招募渠道': 1, '入職渠道': 1,
                       '渠道': 1, '原来源渠道': 1, '中介名称': 1, '派遣公司A（面试渠道）': 1, '派遣公司B（报到渠道）': 1,
                       '中介字段': 1, '中介来源': 1, '中介栏位': 1, '供应商': 1, '来源': 1, '返费': 2,
                       '返费金额': 2, '实返费': 2, '员工返费金额': 2, '员工返费-厂商提供': 2, '打款员工返费金额': 2,
                       '打款员工金额': 2, '中介': 2, '应返费用': 2, '报到返费': 3, '返费价格': 2, '报导返费': 2,
                       '返费情况': 2, '打款员工返费': 2, '应发返费': 2, '000返费': 2, '第一笔返费': 2, '员工返费': 2,
                       '在职打卡25天返费': 2, '18天返费': 2, '25天返费': 2, '30天返费': 2, '45天返费': 2,
                       '应付返费金额': 2, '报道返费价格': 2, '18天返': 2, '返费（给求职者）': 2, '价格（报到返）': 2,
                       '供应•商': 1, '业务所属': 1, '供鹿商': 1, '供应豚': 1, '棋应商': 1, '报到返费  到职日期  离职日期 离职原因': 3,
                       '員工姓名': 0,  '费用': 2, '应支付费用': 2, '入职反费': 2, '报到返 金额': 2,
                       '价格': 2, '实际返钱': 2, '结帐金额': 2, '受款厂商': 1, '原供应商': 1, '管道': 1, '利润': 3, '25天返': 3,
                       '返款金额': 2, '一次返费': 3, '二次返费': 3, '报道返费': 2, '18天': 2, '按照朱巍巍价格': 2, '报到价格': 2,
                       '实际 返钱': 2, '定价': 2, '联系人': 1, '派遣公司A': 1, '應出賬（給面試渠道）': 2, '應出賬': 2, '后期返费': 2,
                       '33天费用': 2, '姓名*': 0, '金額': 2, '派遣別': 1, '勞務公司': 1, '勞務公司名稱': 1, '会员': 0, '派遗公司': 1,
                       '姓名 ▼': 0, '姓名▼': 0, '費用': 2, '入职返': 2, '返款明细': 2, '企业返款明细': 2, '户主': 0, '總計': 2,
                       '员工姓名': 0, '派遣公司名稱': 1, '总价': 2, '與二包商確認金額(二包商提供)': 2, '派遣公司': 1,
                       '来源': 1, '受款廠商': 1, '原派遣公司': 1, '派遣公司A（面試）': 1, '金额': 2, '來源供應商': 1, '管理費': 2,
                       '派遣公司B': 1, '與二包商確認金額': 2, '实际返金': 2, '实际 返銭': 2, '来源渠道': 1, '介绍人': 1, '企业': 1,
                       }

    for name, name_list in table_b_names.items():
        rows, cols = get_sheet_rows_and_cols(ws)

        with open('./1表/处理后/' + name + '-清洗结果.csv', 'a+', newline="", encoding='gb18030') as csv_f:
            csv_writer = csv.writer(csv_f)

            logger.info('开始提取 %s.csv 数据' % name)

            if is_xlsx(ws):
                for i in range(1, rows + 1):
                    write_line = ['' for i in range(7)]
                    line = {}  # {值：字段,...}
                    for j in range(1, cols + 1):
                        # if ws.cell(row=i, column=j).value:
                        value = ws.cell(row=i, column=j).value
                        line[j] = value

                    # if i % 100000:
                    #     print("已处理 %i " % i)

                    # 遍历到表头时，获取表头字段位置，并记录需保存字段位置
                    for k, v in line.items():
                        if v in titles:
                            titles_site = {}
                            keep_index = []

                            for key, value in line.items():
                                if value in titles:
                                    # { 当前位置：写入行位置 }
                                    titles_site[key] = list_index_site[value]

                                    # 获取保留表头索引位置
                                    keep_index.append(int(key))
                            break
                    # 非表头数据，根据记录的表头字段位置匹配，写入CSV文件
                    else:
                        try:
                            for k, v in line.items():
                                if v in name_list:
                                    write_line[4] = ws_name
                                    write_line[5] = i
                                    write_line[6] = str(keep_index)

                                    for item in keep_index:
                                        index = titles_site[item]
                                        write_line[index] = line[item]

                                    if not (write_line[0] == None or write_line[0] == ""):
                                        csv_writer.writerow(write_line)

                                    break

                        except Exception as err:
                            print(i)
                            print(line)
                            print(name_list)
                            print(list_index_site)
                            print(titles_site)
                            print(keep_index)
                            print(write_line)
                            raise err
            else:
                for i in range(rows):
                    write_line = ['' for i in range(7)]
                    line = {}  # {值：字段,...}
                    for j in range(cols):
                        # if ws.cell(row=i, column=j).value:
                        value = ws.cell(i, j).value
                        line[j] = value

                    # if i % 100000:
                    #     print("已处理 %i " % i)

                    # 遍历到表头时，获取表头字段位置，并记录需保存字段位置
                    for k, v in line.items():
                        if v in titles:
                            titles_site = {}
                            keep_index = []

                            for key, value in line.items():
                                if value in titles:
                                    # { 当前位置：写入行位置 }
                                    titles_site[key] = list_index_site[value]

                                    # 获取保留表头索引位置
                                    keep_index.append(int(key))
                            break
                    # 非表头数据，根据记录的表头字段位置匹配，写入CSV文件
                    else:
                        try:
                            for k, v in line.items():
                                if v in name_list:
                                    write_line[4] = ws_name
                                    write_line[5] = i
                                    write_line[6] = str(keep_index)

                                    for item in keep_index:
                                        index = titles_site[item]
                                        write_line[index] = line[item]

                                    if not (write_line[0] == None or write_line[0] == ""):
                                        csv_writer.writerow(write_line)

                                    break

                        except Exception as err:
                            print(i)
                            print(line)
                            print(name_list)
                            print(list_index_site)
                            print(titles_site)
                            print(keep_index)
                            print(write_line)
                            raise err

@run_time
def t():
    path = os.path.abspath('.') + '/1表/'

    # 遍历2表，获取2表中所有姓名字段，生成字典，格式{2表1名：姓名集合, 2表1名：姓名集合} ,使用集合自动去重
    all_tables_b_names = {}
    for table_b_path in os.listdir('./2表劳务公司已排序/'):
        table_b_path_name = os.path.abspath(table_b_path)
        if os.path.isfile(table_b_path_name):
            wb_b = open_workbook(os.path.abspath(table_b_path))
            table_b_names = get_table_b_names(wb_b)

            base_name = os.path.basename(table_b_path).replace('.xlsx', '').replace('.xls', '')

            all_tables_b_names[base_name] = table_b_names
            # print(table_b_names)

    # 生成保存筛选数据的CSV文件，添加表头数据
    for name in all_tables_b_names.keys():
        with open('./1表/处理后/' + name + '-清洗结果.csv', 'w', newline="", encoding='gb18030') as f:
            csv_writer = csv.writer(f)
            csv_writer.writerow(['姓名', '劳务公司', '返费1', '返费2', '数据源', '数据源行', '数据字段'])

    # 遍历1表目录，执行数据清洗
    for file_name in os.listdir('./1表/'):
        file_path = os.path.abspath(file_name)

        if os.path.isfile(file_path):
            # print(save_path)
            if os.path.exists(file_path) and is_xlsx(file_path):
                wb = open_workbook(file_path)
                ws = get_sheets(wb, sheet_name="总表")

                # 数据处理
                clear_content(ws, all_tables_b_names, file_path)
                wb.close()
            else:
                print("未找到文件或类型错误")


if __name__ == '__main__':
    t()

