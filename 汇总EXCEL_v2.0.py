# coding:utf-8
import xlrd
from xlutils.copy import copy
import openpyxl
import datetime
import logging
import os
import multiprocessing
from functools import wraps


logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)  # Log等级总开关
# logging.basicConfig(level=logging.info, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(processName)s - [ %(levelname)s ]：  %(message)s')


def run_time(func):
    """运行耗时装饰器"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        stime = datetime.datetime.now()
        logger.info("开始执行")
        res = func(*args, **kwargs)
        etime = datetime.datetime.now()
        cost = str(etime - stime).split(":")
        logger.info("执行耗时：[ " + cost[0] + '时' + cost[1] + "分" + cost[2][0:5] + "秒" + " ]")
        return res
    return wrapper


def is_xlsx(workbook):
    return True if str(type(workbook)).split("'")[1].split(".")[0] == 'openpyxl' else False


def get_sheets(workbook):
    """获取工作表中所有工作簿，返回工作簿对象列表"""
    try:
        table_a_sheets = []
        if is_xlsx(workbook):
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                if not ws.max_row == 0:
                    table_a_sheets.append(workbook[sheet])
        else:
            for sheet in workbook.sheet_names():
                ws = workbook.sheet_by_name(sheet)
                if not ws.nrows == 0:
                    table_a_sheets.append(ws)
        return table_a_sheets
    except AttributeError as e:
        logger.error("获取工作簿对象列表错误")
        raise e


def get_sheet_rows_and_cols(sheet):
    """获取工作簿的行数和列数，返回工作簿对象的行、列的数组"""
    try:
        sheet_type = True if str(type(sheet)).split("'")[1].split(".")[0] == 'openpyxl' else False
        if sheet_type:
            return sheet.max_row, sheet.max_column
        else:
            return sheet.nrows, sheet.ncols
    except AttributeError as e:
        logger.error("获取行数、列数错误，参数错误")
        raise e


def get_table_b_namess_and_data(sheet, table_b_index_column):
    """获取sheet的匹配名称列表，及详细数据列表"""
    b_rows, b_cols = get_sheet_rows_and_cols(sheet)

    b_names = set()  # 创建一个集合（去重，集合不会重复）
    b_all_data = {}  # 创建一个字典存放所有B表中的去重数据

    if is_xlsx(sheet):
        # openpyxl 单元格索引从1开始
        for row_index in range(1, b_rows + 1):
            if sheet.cell(row=row_index, column=table_b_index_column).value:
                b_names.add(sheet.cell(row=row_index, column=table_b_index_column).value)
                temp_list = []
                for col_index in range(1, b_cols + 1):
                    temp_list.append(sheet.cell(row=row_index, column=col_index).value)
                b_all_data[sheet.cell(row=row_index, column=table_b_index_column).value] = temp_list
    else:
        # xlrd 单元格索引从0开始
        for row_index in range(b_rows):
            if sheet.cell(row_index, table_b_index_column - 1).value:
                b_names.add(sheet.cell(row_index, table_b_index_column - 1).value)
                temp_list = []
                for col_index in range(b_cols):
                    temp_list.append(sheet.cell(row_index, col_index).value)
                b_all_data[sheet.cell(row_index, table_b_index_column - 1).value] = temp_list
    return b_names, b_all_data


def merge_table_a_data(ws_a, ws_b, a_index=None, b_index=None, nws=None):
    """表A数据合并"""
    a_rows, a_cols = get_sheet_rows_and_cols(ws_a)
    b_rows, b_cols = get_sheet_rows_and_cols(ws_b)

    max_cols = a_cols + b_cols

    # 获取表B数据
    b_names, b_all_data = get_table_b_namess_and_data(ws_b, b_index)
    if is_xlsx(ws_a):
        # 判断name值是否存在，存在则在A表该列后面追加B表name值一致的行
        for row_index in range(1, a_rows + 1):  # 遍历A表每一行
            if ws_a.cell(row=row_index, column=a_index).value in b_names:  # 判断如果name在集合中
                for a_col, b_col in zip(range(a_cols + 1, max_cols + 1),
                                        range(1, b_cols + 1)):  # 产生表A最大列 ~ 最大列+表B最大列数 迭代对象
                    # 设：表A列数为10，表B列数为5，即产生：11~15 ，在11~15列插入表B 列1~列5的数据
                    ws_a.cell(row=row_index, column=a_col).value = \
                        b_all_data[ws_a.cell(row=row_index, column=a_index).value][b_col - 1]
    else:
        for row_index in range(a_rows):
            if ws_a.cell(row_index, a_index - 1).value in b_names:
                for a_col, b_col in zip(range(a_cols, max_cols), range(b_cols)):
                    nws.write(row_index, a_col,
                                 b_all_data[ws_a.cell(row_index, a_index - 1).value][b_col])


def save_work_book(workbook, save_file_path):
    logger.info("开始保存数据")
    workbook.save(save_file_path)
    logger.info("保存数据成功")


# @run_time
def run(source_dict):
    """程序执行点"""
    wb_a_name = source_dict['tableA']
    wb_b_names = source_dict['tableB']
    table_a_index_column = source_dict['tableA_index_column']
    table_b_index_column = source_dict['tableB_index_column']
    save_file_path = os.path.dirname(wb_a_name) + "/" + os.path.basename(wb_a_name).split(".")[0] + "_合并." + \
                     os.path.basename(wb_a_name).split(".")[1]

    if not os.path.exists(wb_a_name):
        logger.error("未找到B表，请确认")
        return 1

    if not os.path.exists(wb_b_names):
        logger.error("未找到B表，请确认")
        return 1

    wb_a_type = True if wb_a_name.endswith(".xlsx") else False
    wb_b_type = True if wb_b_names.endswith(".xlsx") else False

    logger.info("加载表格")
    if wb_a_type:
        wb_a = openpyxl.load_workbook(wb_a_name)  # 读取xlsx格式的A表
    else:
        wb_a = xlrd.open_workbook(wb_a_name)  # 读取xls格式的A表

    if wb_b_type:
        wb_b = openpyxl.load_workbook(wb_b_names)  # 读取xlsx格式的B表
    else:
        wb_b = xlrd.open_workbook(wb_b_names)  # 读取xls格式的B表

    # 获取非空的工作簿列表
    ws_a_list = get_sheets(wb_a)
    ws_b_list = get_sheets(wb_b)

    logger.info("正在进行数据处理")
    if is_xlsx(wb_a):
        for ws_a in ws_a_list:
            for ws_b in ws_b_list:
                merge_table_a_data(ws_a, ws_b, a_index=table_a_index_column, b_index=table_b_index_column)
        save_work_book(wb_a, save_file_path)
    else:
        for ws_a in ws_a_list:
            for ws_b in ws_b_list:
                new_wb = copy(wb_a)
                new_ws = new_wb.get_sheet(ws_a.name)
                merge_table_a_data(ws_a, ws_b,a_index=table_a_index_column, b_index=table_b_index_column, nws=new_ws)
        save_work_book(new_wb, save_file_path)


if __name__ == '__main__':
    data_source = [
        {
            "tableA": 'C:/Users/lqcma/Desktop/excel/B表1.xlsx',
            "tableB":'C:/Users/lqcma/Desktop/excel/王孝华.xls',
            "tableA_index_column": 2,   # 表A条件匹配列,从1开始计数
            "tableB_index_column": 7,   # 表B条件匹配列,从1开始计数
         },
        {
            "tableA":'C:/Users/lqcma/Desktop/excel/王孝华.xls',
            "tableB":'C:/Users/lqcma/Desktop/excel/B表1.xlsx',

            "tableA_index_column": 7,   # 表A条件匹配列,从1开始计数
            "tableB_index_column": 2,   # 表B条件匹配列,从1开始计数
        },
    ]

    if len(data_source) < 1:
        logger.error("数据源为空，请配置数据源，程序结束")

    elif len(data_source) == 1:
        logger.info("数据源数量为 1 开始处理")
        run(data_source[0])
    else:
        # 大数据量Excel会占用极大内存，多进程并发，有可能会造成MemoryError,如MemoryError请使用单进程
        logger.info("数据源数量为 %i 可配置多进程处理，请注意内存消耗，数据量大的Excel容易造成MemoryError" % len(data_source))
        logger.info("当前CPU核心数为 %i 请根据需要配置并发数" % multiprocessing.cpu_count())
        # p = multiprocessing.Pool(multiprocessing.cpu_count())
        p = multiprocessing.Pool(1)  # 设置参数1即为单线程
        for i, source in zip(range(len(data_source)), data_source):
            p.apply_async(run, args=(source,))
        p.close()
        p.join()

        # 以下为直接创建并运行所有进程
        # process_list = []
        # for i, s in enumerate(data_source):
        #     i = multiprocessing.Process(target=run, args=(s,))
        #     process_list.append(i)
        #     i.start()
        #
        # for i in process_list:
        #     i.join()

        logger.info("程序执行完毕")




    # main()

